"""
================================================
  PharmaTrack — Excel Import Script
================================================

HOW TO USE:
1. Download your Google Sheet as Excel (.xlsx)
2. Put the file in this same pharma_app folder
3. Rename it to: patients_data.xlsx
4. Double-click IMPORT_DATA.bat (or run: python import_excel.py)
5. Done! Open http://localhost:5000 to see all your patients

Your existing app data will be BACKED UP before import.
================================================
"""

import openpyxl
import json
import os
import uuid
import shutil
from datetime import datetime, date

# ── CONFIG ──────────────────────────────────────────
EXCEL_FILE  = "patients_data.xlsx"
DATA_FILE   = os.path.join("data", "patients.json")
BACKUP_FILE = os.path.join("data", "patients_backup.json")

# Column mapping — matches your Google Sheet exactly
# (Column letters from your sheet → row[index], 0-based)
# B=NAME, C=UHID, D=Old/New, E=AGE, F=SEX,
# G=DOV, H=PHONE, I=ADDRESS, J=CURRENT STATUS,
# K=DIAGNOSIS, L=FOLLOW UP REMARKS, M=NEXT FOLLOW UP DATE
COL_NAME        = 0   # B
COL_UHID        = 1   # C
COL_OLD_NEW     = 2   # D
COL_AGE         = 3   # E
COL_SEX         = 4   # F
COL_DOV         = 5   # G  (Date of Visit)
COL_PHONE       = 6   # H
COL_ADDRESS     = 7   # I
COL_STATUS      = 8   # J  (Current Status)
COL_DIAGNOSIS   = 9   # K
COL_REMARKS     = 10  # L  (Follow Up Visit Remarks)
COL_FOLLOWUP    = 11  # M  (Next Follow Up Visit Date)
# ────────────────────────────────────────────────────

def parse_date(val):
    """Try to parse various date formats into YYYY-MM-DD"""
    if not val:
        return ""
    if isinstance(val, (datetime,)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ['none', 'nan', '-', '']:
        return ""
    # Try common formats
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y",
                "%d.%m.%Y", "%d.%m.%y", "%d-%m-%y", "%d/%m/%y",
                "%d.%m.%Y", "%-d.%-m.%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    # Handle formats like "1.08.25" or "2.08.25"
    parts = s.replace('/', '.').replace('-', '.').split('.')
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 2:
            y = "20" + y
        try:
            return datetime.strptime(f"{d}.{m}.{y}", "%d.%m.%Y").strftime("%Y-%m-%d")
        except:
            pass
    return s  # Return as-is if nothing works

def clean(val):
    """Clean a cell value to string"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ['none', 'nan']:
        return ""
    return s

def determine_status(followup_date_str, current_status_str):
    """Decide if patient goes to active or resolved based on their data"""
    status_lower = current_status_str.lower()

    # Keywords that suggest case is closed/resolved
    resolved_keywords = [
        'already', 'not interested', 'expired', 'deceased',
        'closed', 'resolved', 'done', 'completed', 'discharged'
    ]
    for kw in resolved_keywords:
        if kw in status_lower:
            return 'resolved'

    # If there's a future follow-up date, it's active
    if followup_date_str:
        try:
            fdate = datetime.strptime(followup_date_str, "%Y-%m-%d").date()
            return 'active'
        except:
            pass

    # Default to active
    return 'active'

def import_patients():
    print("\n" + "="*55)
    print("  PharmaTrack — Excel Import Tool")
    print("="*55)

    # Check Excel file exists
    if not os.path.exists(EXCEL_FILE):
        print(f"\n❌ ERROR: Could not find '{EXCEL_FILE}'")
        print(f"   Make sure you:")
        print(f"   1. Downloaded your Google Sheet as .xlsx")
        print(f"   2. Named it exactly: patients_data.xlsx")
        print(f"   3. Put it in this folder: {os.path.abspath('.')}")
        input("\nPress Enter to exit...")
        return

    # Backup existing data
    os.makedirs("data", exist_ok=True)
    if os.path.exists(DATA_FILE):
        shutil.copy(DATA_FILE, BACKUP_FILE)
        print(f"\n✅ Backup saved: {BACKUP_FILE}")

    # Load Excel
    print(f"\n📂 Reading: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    print(f"   Sheet: '{ws.title}' — {ws.max_row} rows found")

    active_patients = []
    resolved_patients = []
    skipped = 0
    imported = 0

    # Find the header row (look for 'NAME' in column B)
    start_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == 'NAME':
                start_row = cell.row + 1
                print(f"   Header found at row {cell.row}, importing from row {start_row}")
                break
        if start_row:
            break

    if not start_row:
        print("   ⚠️  Could not find header row — starting from row 4 (default)")
        start_row = 4

    # Process each row
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # Get values using column positions (offset from col B = index 1)
        vals = list(row)
        if len(vals) < 2:
            skipped += 1
            continue

        # Offset: sheet col A is index 0, col B is index 1
        offset = 1
        name = clean(vals[offset + COL_NAME]) if len(vals) > offset + COL_NAME else ""

        # Skip empty rows
        if not name or name.upper() in ['NAME', 'S.NO', '']:
            skipped += 1
            continue

        uhid        = clean(vals[offset + COL_UHID])      if len(vals) > offset + COL_UHID      else ""
        old_new     = clean(vals[offset + COL_OLD_NEW])   if len(vals) > offset + COL_OLD_NEW   else ""
        age         = clean(vals[offset + COL_AGE])       if len(vals) > offset + COL_AGE       else ""
        sex         = clean(vals[offset + COL_SEX])       if len(vals) > offset + COL_SEX       else ""
        dov         = parse_date(vals[offset + COL_DOV])  if len(vals) > offset + COL_DOV       else ""
        phone       = clean(vals[offset + COL_PHONE])     if len(vals) > offset + COL_PHONE     else ""
        address     = clean(vals[offset + COL_ADDRESS])   if len(vals) > offset + COL_ADDRESS   else ""
        curr_status = clean(vals[offset + COL_STATUS])    if len(vals) > offset + COL_STATUS    else ""
        diagnosis   = clean(vals[offset + COL_DIAGNOSIS]) if len(vals) > offset + COL_DIAGNOSIS else ""
        remarks     = clean(vals[offset + COL_REMARKS])   if len(vals) > offset + COL_REMARKS   else ""
        followup    = parse_date(vals[offset + COL_FOLLOWUP]) if len(vals) > offset + COL_FOLLOWUP else ""

        patient = {
            "id":             str(uuid.uuid4()),
            "name":           name,
            "uhid":           uhid,
            "old_new":        old_new,
            "age":            age,
            "sex":            sex,
            "phone":          phone,
            "doctor":         "",        # Not in your sheet — can fill later
            "area":           address,
            "product":        "",        # Not in your sheet — can fill later
            "diagnosis":      diagnosis,
            "medicine":       "",        # Not in your sheet — can fill later
            "current_status": curr_status,
            "rep":            "",        # Not in your sheet — can fill later
            "notes":          remarks,
            "visit_date":     dov,
            "followup_date":  followup,
            "added_on":       date.today().strftime("%Y-%m-%d")
        }

        bucket = determine_status(followup, curr_status)
        if bucket == 'resolved':
            patient['resolved_on'] = date.today().strftime("%Y-%m-%d")
            resolved_patients.append(patient)
        else:
            active_patients.append(patient)

        imported += 1

    # Save to app data file
    data = {"active": active_patients, "resolved": resolved_patients}
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\n✅ IMPORT COMPLETE!")
    print(f"   Total imported : {imported} patients")
    print(f"   Active         : {len(active_patients)} patients")
    print(f"   Resolved       : {len(resolved_patients)} patients")
    print(f"   Skipped (empty): {skipped} rows")
    print(f"\n🌐 Now open your browser and go to:")
    print(f"   http://localhost:5000")
    print(f"\n   (Start the app first with START_APP.bat if not running)")
    print("\n" + "="*55)
    input("\nPress Enter to exit...")

if __name__ == "__main__":
    import_patients()
